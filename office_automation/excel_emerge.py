# encoding:utf-8
import os
import glob
import openpyxl


# 获取指定目录下所有的 xlsx 文件
xlsx_files = glob.glob(os.path.join(os.getcwd()+'/employee', '*.xlsx'))

# 创建一个新的工作簿
wb_new = openpyxl.Workbook()
ws_new = wb_new.active
ws_new.title = 'merge'
is_first_file = True

# 遍历所有的Excel文件
for filename in xlsx_files:
    wb = openpyxl.load_workbook(filename)
    # 获取每个Excel文件中活跃的工作表
    ws = wb.active
    # 按行获取所有单元格
    if is_first_file:
        for row in ws.iter_rows(min_row=1):
            values = []
            for cell in row:
                values.append(cell.value)
            # values = [cell.value for cell in row]
            # 向表格末尾添加数据
            ws_new.append(values)
            is_first_file = False
    else:
        for row in ws.iter_rows(min_row=2):
            values = []
            for cell in row:
                values.append(cell.value)
            # values = [cell.value for cell in row]
            # 向表格末尾添加数据
            ws_new.append(values)

wb_new.save('merge.xlsx')
wb_new.close()
